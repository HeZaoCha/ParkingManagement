"""
排班管理视图

提供排班表的Excel导入导出功能。

Author: HeZaoCha
Created: 2025-12-11
Version: 1.1.0
"""
from io import BytesIO
from typing import Optional

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from parking.decorators import staff_member_required
from parking.models import ParkingLot
from parking.user_models import StaffSchedule


@staff_member_required
@require_http_methods(['GET'])
def schedule_template_download(request):
    """下载排班表模板Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '排班表模板'
    
    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置表头
    headers = ['用户名', '停车场名称', '星期', '开始时间', '结束时间', '是否启用']
    ws.append(headers)
    
    # 应用表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 添加示例数据
    examples = [
        ['staff1', '早点喝茶停车场', '周一', '08:00', '16:00', '是'],
        ['staff2', '早点喝茶停车场', '周二', '08:00', '16:00', '是'],
    ]
    for row in examples:
        ws.append(row)
    
    # 添加说明
    ws.append([])
    ws.append(['说明：'])
    ws.append(['1. 星期：周一、周二、周三、周四、周五、周六、周日'])
    ws.append(['2. 时间格式：HH:MM，如 08:00'])
    ws.append(['3. 是否启用：是/否'])
    ws.append(['4. 用户名必须是已存在的工作人员账号'])
    ws.append(['5. 停车场名称必须与系统中已存在的停车场名称完全一致'])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # 生成响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="排班表模板.xlsx"'
    
    wb.save(response)
    return response


@staff_member_required
@require_http_methods(['POST'])
def schedule_upload(request):
    """上传并解析排班表Excel"""
    if 'file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'message': '请选择要上传的文件'
        }, status=400)
    
    file = request.FILES['file']
    
    # 验证文件类型
    if not file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'message': '只支持Excel文件（.xlsx, .xls）'
        }, status=400)
    
    try:
        # 读取Excel文件
        wb = load_workbook(filename=BytesIO(file.read()))
        ws = wb.active
        
        # 解析数据（跳过表头）
        success_count = 0
        error_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # 跳过空行
                continue
            
            try:
                username = str(row[0]).strip()
                lot_name = str(row[1]).strip()
                weekday_str = str(row[2]).strip()
                start_time_str = str(row[3]).strip()
                end_time_str = str(row[4]).strip()
                is_active_str = str(row[5]).strip() if len(row) > 5 else '是'
                
                # 验证用户
                # 优化：使用select_related预加载groups
                from django.contrib.auth.models import User
                try:
                    user = User.objects.select_related().prefetch_related('groups').get(username=username)
                    if not user.groups.filter(name__in=['Staff', 'Admin']).exists():
                        raise ValueError(f'用户 {username} 不是工作人员')
                except User.DoesNotExist:
                    raise ValueError(f'用户 {username} 不存在')
                
                # 验证停车场
                try:
                    parking_lot = ParkingLot.objects.get(name=lot_name)
                except ParkingLot.DoesNotExist:
                    raise ValueError(f'停车场 {lot_name} 不存在')
                
                # 解析星期
                weekday_map = {
                    '周一': 0, '周二': 1, '周三': 2, '周四': 3,
                    '周五': 4, '周六': 5, '周日': 6,
                    '星期一': 0, '星期二': 1, '星期三': 2, '星期四': 3,
                    '星期五': 4, '星期六': 5, '星期日': 6,
                }
                weekday = weekday_map.get(weekday_str)
                if weekday is None:
                    raise ValueError(f'无效的星期: {weekday_str}')
                
                # 解析时间
                from datetime import datetime
                try:
                    start_time = datetime.strptime(start_time_str, '%H:%M').time()
                    end_time = datetime.strptime(end_time_str, '%H:%M').time()
                except ValueError:
                    raise ValueError(f'时间格式错误: {start_time_str} 或 {end_time_str}，应为 HH:MM')
                
                # 解析是否启用
                is_active = is_active_str.lower() in ['是', 'yes', 'true', '1', '启用']
                
                # 创建或更新排班
                schedule, created = StaffSchedule.objects.update_or_create(
                    user=user,
                    parking_lot=parking_lot,
                    weekday=weekday,
                    defaults={
                        'start_time': start_time,
                        'end_time': end_time,
                        'is_active': is_active,
                    }
                )
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'第{row_idx}行: {str(e)}')
        
        return JsonResponse({
            'success': True,
            'message': f'导入完成：成功 {success_count} 条，失败 {error_count} 条',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]  # 最多返回10个错误
        })
        
    except Exception as e:
        logger.exception('排班表导入失败')
        return JsonResponse({
            'success': False,
            'message': f'导入失败: {str(e)}'
        }, status=500)


@staff_member_required
@require_http_methods(['GET'])
def schedule_list(request):
    """排班表列表"""
    schedules = StaffSchedule.objects.select_related(
        'user', 'parking_lot'
    ).filter(is_active=True).order_by('parking_lot', 'weekday', 'start_time')
    
    # 按停车场分组
    grouped = {}
    for schedule in schedules:
        lot_name = schedule.parking_lot.name
        if lot_name not in grouped:
            grouped[lot_name] = []
        grouped[lot_name].append(schedule)
    
    context = {
        'schedules': schedules,
        'grouped_schedules': grouped,
    }
    return render(request, 'admin/schedule/list.html', context)

