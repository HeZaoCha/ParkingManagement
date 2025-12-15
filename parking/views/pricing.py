"""
费率管理视图

提供费率模板的创建、编辑、删除等功能。

Author: HeZaoCha
Created: 2025-12-11
Version: 1.1.0
"""

import json
import zipfile
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_http_methods
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from parking.decorators import staff_member_required
from parking.pricing_models import (
    MonthYearRate,
    OvertimeRate,
    ParkingLotPricing,
    PricingRule,
    PricingTemplate,
)


@staff_member_required
@require_http_methods(["GET"])
def pricing_template_list(request):
    """
    费率模板列表
    
    注意：不缓存此页面，因为管理后台数据变化频繁，需要实时显示最新数据。
    """
    templates = PricingTemplate.objects.all().prefetch_related("rules").order_by("-created_at")

    context = {
        "templates": templates,
    }
    return render(request, "admin/pricing/template_list.html", context)


@staff_member_required
@require_http_methods(["GET", "POST"])
def pricing_template_edit(request, template_id=None):
    """创建/编辑费率模板"""
    template = None
    if template_id:
        template = get_object_or_404(PricingTemplate, id=template_id)

    if request.method == "POST":
        try:
            data = (
                json.loads(request.body)
                if request.content_type == "application/json"
                else request.POST
            )

            if template:
                template.name = data.get("name", template.name)
                template.description = data.get("description", "")
                template.free_minutes = int(data.get("free_minutes", 15))
                template.daily_max_fee = data.get("daily_max_fee") or None
            else:
                template = PricingTemplate.objects.create(
                    name=data.get("name"),
                    description=data.get("description", ""),
                    free_minutes=int(data.get("free_minutes", 15)),
                    daily_max_fee=data.get("daily_max_fee") or None,
                )

            # 更新费率规则
            rules_data = data.get("rules", [])
            if isinstance(rules_data, str):
                rules_data = json.loads(rules_data)

            # 删除旧规则
            if template_id:
                template.rules.all().delete()
                template.month_year_rates.all().delete()
                template.overtime_rates.all().delete()

            # 创建新规则
            for idx, rule_data in enumerate(rules_data):
                PricingRule.objects.create(
                    template=template,
                    start_minutes=int(rule_data.get("start_minutes", 0)),
                    end_minutes=int(rule_data.get("end_minutes", 0))
                    if rule_data.get("end_minutes")
                    else None,
                    rate_per_hour=rule_data.get("rate_per_hour"),
                    vehicle_type=rule_data.get("vehicle_type", "all"),
                    order=idx,
                )

            # 更新包月/包年费率
            month_year_data = data.get("month_year_rates", [])
            if isinstance(month_year_data, str):
                month_year_data = json.loads(month_year_data)

            for rate_data in month_year_data:
                MonthYearRate.objects.create(
                    template=template,
                    rate_type=rate_data.get("rate_type"),
                    price=rate_data.get("price"),
                    vehicle_type=rate_data.get("vehicle_type", "all"),
                    description=rate_data.get("description", ""),
                    is_active=rate_data.get("is_active", True),
                )

            # 更新超时收费
            overtime_data = data.get("overtime_rates", [])
            if isinstance(overtime_data, str):
                overtime_data = json.loads(overtime_data)

            for overtime_rate_data in overtime_data:
                OvertimeRate.objects.create(
                    template=template,
                    overtime_fee=overtime_rate_data.get("overtime_fee"),
                    overtime_start_hours=overtime_rate_data.get("overtime_start_hours", 24),
                    vehicle_type=overtime_rate_data.get("vehicle_type", "all"),
                    description=overtime_rate_data.get("description", ""),
                    is_active=overtime_rate_data.get("is_active", True),
                )

            template.save()

            return JsonResponse(
                {"success": True, "message": "模板保存成功", "template_id": template.id}
            )

        except Exception as e:
            logger.exception("保存费率模板失败")
            return JsonResponse({"success": False, "message": f"保存失败: {str(e)}"}, status=500)

    # GET请求：显示编辑页面
    # 优化：使用prefetch_related预加载规则
    if template:
        template = (
            PricingTemplate.objects.prefetch_related("rules", "month_year_rates", "overtime_rates")
            .get(id=template.id)
        )
        rules = list(template.rules.all().order_by("order", "start_minutes"))
        month_year_rates = list(template.month_year_rates.all())
        overtime_rates = list(template.overtime_rates.all())
    else:
        rules = []
        month_year_rates = []
        overtime_rates = []

    context = {
        "template": template,
        "rules": rules,
        "month_year_rates": month_year_rates,
        "overtime_rates": overtime_rates,
    }
    return render(request, "admin/pricing/template_edit.html", context)


@staff_member_required
@require_http_methods(["POST"])
def pricing_template_delete(request, template_id):
    """删除费率模板"""
    template = get_object_or_404(PricingTemplate, id=template_id)

    # 检查是否被使用
    if template.parking_lots.exists():
        return JsonResponse({"success": False, "message": "该模板正在被使用，无法删除"}, status=400)

    template_name = template.name
    template.delete()

    logger.info(f"成功删除费率模板: {template_name}")

    return JsonResponse({"success": True, "message": "模板删除成功"})


@staff_member_required
@require_http_methods(["POST"])
def pricing_template_copy(request, template_id):
    """拷贝费率模板"""
    source_template = get_object_or_404(PricingTemplate, id=template_id)

    try:
        data = (
            json.loads(request.body)
            if request.content_type == "application/json"
            else request.POST
        )
        new_name = data.get("name", f"{source_template.name} (副本)")

        # 检查名称是否已存在
        if PricingTemplate.objects.filter(name=new_name).exists():
            return JsonResponse(
                {"success": False, "message": f"模板名称 '{new_name}' 已存在"}, status=400
            )

        # 创建新模板
        new_template = PricingTemplate.objects.create(
            name=new_name,
            description=source_template.description,
            free_minutes=source_template.free_minutes,
            daily_max_fee=source_template.daily_max_fee,
            is_active=source_template.is_active,
        )

        # 拷贝费率规则
        for rule in source_template.rules.all():
            PricingRule.objects.create(
                template=new_template,
                start_minutes=rule.start_minutes,
                end_minutes=rule.end_minutes,
                rate_per_hour=rule.rate_per_hour,
                vehicle_type=rule.vehicle_type,
                order=rule.order,
            )

        # 拷贝包月/包年费率
        for rate in source_template.month_year_rates.all():
            MonthYearRate.objects.create(
                template=new_template,
                rate_type=rate.rate_type,
                price=rate.price,
                vehicle_type=rate.vehicle_type,
                description=rate.description,
                is_active=rate.is_active,
            )

        # 拷贝超时收费
        for overtime in source_template.overtime_rates.all():
            OvertimeRate.objects.create(
                template=new_template,
                overtime_fee=overtime.overtime_fee,
                overtime_start_hours=overtime.overtime_start_hours,
                vehicle_type=overtime.vehicle_type,
                description=overtime.description,
                is_active=overtime.is_active,
            )

        logger.info(f"成功拷贝费率模板: {source_template.name} -> {new_template.name}")

        return JsonResponse(
            {
                "success": True,
                "message": "模板拷贝成功",
                "template_id": new_template.id,
                "redirect_url": f"/parking/manage/pricing/templates/{new_template.id}/",
            }
        )

    except Exception as e:
        logger.exception("拷贝费率模板失败")
        return JsonResponse({"success": False, "message": f"拷贝失败: {str(e)}"}, status=500)


@staff_member_required
@require_http_methods(["GET"])
def pricing_template_create_from(request, template_id=None):
    """基于现有模板创建新模板（重定向到编辑页面）"""
    source_template = None
    if template_id:
        source_template = get_object_or_404(
            PricingTemplate.objects.prefetch_related("rules", "month_year_rates", "overtime_rates"),
            id=template_id,
        )

    # 创建新模板（基于源模板或空白）
    if source_template:
        new_template = PricingTemplate.objects.create(
            name=f"{source_template.name} (新模板)",
            description=source_template.description,
            free_minutes=source_template.free_minutes,
            daily_max_fee=source_template.daily_max_fee,
            is_active=True,
        )

        # 拷贝费率规则
        for rule in source_template.rules.all():
            PricingRule.objects.create(
                template=new_template,
                start_minutes=rule.start_minutes,
                end_minutes=rule.end_minutes,
                rate_per_hour=rule.rate_per_hour,
                vehicle_type=rule.vehicle_type,
                order=rule.order,
            )

        # 拷贝包月/包年费率
        for rate in source_template.month_year_rates.all():
            MonthYearRate.objects.create(
                template=new_template,
                rate_type=rate.rate_type,
                price=rate.price,
                vehicle_type=rate.vehicle_type,
                description=rate.description,
                is_active=rate.is_active,
            )

        # 拷贝超时收费
        for overtime in source_template.overtime_rates.all():
            OvertimeRate.objects.create(
                template=new_template,
                overtime_fee=overtime.overtime_fee,
                overtime_start_hours=overtime.overtime_start_hours,
                vehicle_type=overtime.vehicle_type,
                description=overtime.description,
                is_active=overtime.is_active,
            )

        logger.info(f"基于模板 '{source_template.name}' 创建新模板: {new_template.name}")
    else:
        # 创建空白模板
        new_template = PricingTemplate.objects.create(
            name="新费率模板",
            description="",
            free_minutes=15,
            daily_max_fee=None,
            is_active=True,
        )

    # 重定向到编辑页面
    from django.shortcuts import redirect

    return redirect("parking:admin_pricing_template_edit", template_id=new_template.id)


@staff_member_required
@require_http_methods(["GET", "POST"])
def parking_lot_pricing_edit(request, lot_id):
    """停车场费率配置"""
    from parking.models import ParkingLot

    parking_lot = get_object_or_404(ParkingLot, id=lot_id)
    pricing_config, created = ParkingLotPricing.objects.get_or_create(parking_lot=parking_lot)

    if request.method == "POST":
        try:
            data = (
                json.loads(request.body)
                if request.content_type == "application/json"
                else request.POST
            )

            pricing_config.charge_type = data.get("charge_type", "fixed")
            template_id = data.get("template_id")

            if template_id:
                template = PricingTemplate.objects.get(id=template_id)
                pricing_config.template = template
            else:
                pricing_config.template = None

            # 使用match/case优化（Python 3.10+特性）
            match pricing_config.charge_type:
                case "fixed":
                    pricing_config.hourly_rate = data.get("hourly_rate")
                    pricing_config.free_minutes = None
                    pricing_config.daily_max_fee = None
                    pricing_config.custom_rules = []
                case "tiered":
                    pricing_config.hourly_rate = None
                    pricing_config.free_minutes = int(data.get("free_minutes", 15))
                    pricing_config.daily_max_fee = data.get("daily_max_fee") or None
                    # 自定义规则（仅在无模板时使用）
                    if not pricing_config.template:
                        custom_rules = data.get("custom_rules", [])
                        if isinstance(custom_rules, str):
                            custom_rules = json.loads(custom_rules)
                        pricing_config.custom_rules = custom_rules
                case _:
                    pass  # 保持原值

            pricing_config.save()

            return JsonResponse({"success": True, "message": "费率配置保存成功"})

        except Exception as e:
            logger.exception("保存费率配置失败")
            return JsonResponse({"success": False, "message": f"保存失败: {str(e)}"}, status=500)

    # GET请求：显示配置页面
    # 优化：使用prefetch_related预加载模板规则
    templates = PricingTemplate.objects.filter(is_active=True).prefetch_related("rules")

    # 优化：如果pricing_config有template，预加载其规则
    if pricing_config.template_id:
        pricing_config = (
            ParkingLotPricing.objects.select_related("template")
            .prefetch_related("template__rules")
            .get(id=pricing_config.id)
        )

    context = {
        "parking_lot": parking_lot,
        "pricing_config": pricing_config,
        "templates": templates,
    }
    return render(request, "admin/parking_lot/pricing_edit.html", context)


@staff_member_required
@require_http_methods(["POST"])
def pricing_preview(request):
    """费率预览API"""
    try:
        data = (
            json.loads(request.body) if request.content_type == "application/json" else request.POST
        )

        duration_minutes = int(data.get("duration_minutes", 0))
        charge_type = data.get("charge_type", "fixed")
        lot_id = data.get("lot_id")

        if not lot_id:
            return JsonResponse({"success": False, "message": "请提供停车场ID"}, status=400)

        from parking.models import ParkingLot, ParkingRecord, ParkingSpace, Vehicle
        from datetime import timedelta
        from django.utils import timezone

        parking_lot = get_object_or_404(ParkingLot, id=lot_id)

        try:
            pricing_config = parking_lot.pricing_config
        except AttributeError:
            pricing_config = None

        if duration_minutes <= 0:
            return JsonResponse({"success": True, "fee": "0.00", "breakdown": []})

        # 计算费用（使用match/case优化）
        match charge_type:
            case "fixed":
                hourly_rate = float(data.get("hourly_rate", parking_lot.hourly_rate))
                free_minutes = 15

                if duration_minutes <= free_minutes:
                    fee = 0.00
                    breakdown = [f"前{free_minutes}分钟免费"]
                else:
                    billable_minutes = duration_minutes - free_minutes
                    hours = (billable_minutes + 59) // 60  # 向上取整
                    fee = hours * hourly_rate
                    breakdown = [
                        f"前{free_minutes}分钟免费",
                        f"计费时长：{billable_minutes}分钟（按{hours}小时计费）",
                        f"费率：¥{hourly_rate:.2f}/小时",
                        f"费用：{hours} × ¥{hourly_rate:.2f} = ¥{fee:.2f}",
                    ]
            case "tiered":
                # 阶梯收费
                # 检查是否有配置或模板
                template_id = data.get("template_id")
                if template_id:
                    template = PricingTemplate.objects.get(id=template_id)
                    free_minutes = template.free_minutes
                elif pricing_config and pricing_config.charge_type == "tiered":
                    free_minutes = pricing_config.get_free_minutes()
                else:
                    free_minutes = int(data.get("free_minutes", 15))

                daily_max_fee = float(data.get("daily_max_fee", 0)) or None

                if duration_minutes <= free_minutes:
                    fee = 0.00
                    breakdown = [f"前{free_minutes}分钟免费"]
                else:
                    # 创建临时记录来计算费用
                    temp_vehicle, _ = Vehicle.objects.get_or_create(
                        license_plate="TEMP_PREVIEW", defaults={"vehicle_type": "car"}
                    )
                    temp_space = ParkingSpace.objects.filter(
                        parking_lot=parking_lot, is_occupied=False
                    ).first()

                    if temp_space:
                        # 临时创建费率配置（如果不存在）
                        if not pricing_config:
                            from parking.pricing_models import ParkingLotPricing

                            pricing_config, _ = ParkingLotPricing.objects.get_or_create(
                                parking_lot=parking_lot, defaults={"charge_type": "tiered"}
                            )

                        if template_id:
                            pricing_config.template_id = template_id
                        pricing_config.charge_type = "tiered"
                        pricing_config.free_minutes = free_minutes
                        pricing_config.daily_max_fee = daily_max_fee
                        pricing_config.save()

                        temp_record = ParkingRecord.objects.create(
                            vehicle=temp_vehicle,
                            parking_space=temp_space,
                            entry_time=timezone.now() - timedelta(minutes=duration_minutes),
                            exit_time=timezone.now(),
                        )
                        fee = float(temp_record.calculate_fee())
                        temp_record.delete()

                        breakdown = [f"前{free_minutes}分钟免费"]
                        breakdown.append(f"计费时长：{duration_minutes - free_minutes}分钟")

                        # 获取规则详情
                        effective_rules = pricing_config.get_effective_rules()
                        if effective_rules:
                            breakdown.append("阶梯规则：")
                            for rule in effective_rules:
                                start = rule.get("start_minutes", 0)
                                end = rule.get("end_minutes", "∞")
                                rate = rule.get("rate_per_hour", 0)
                                breakdown.append(f"  {start}-{end}分钟：¥{rate:.2f}/小时")

                        if daily_max_fee and fee > daily_max_fee:
                            breakdown.append(f"超过每日上限¥{daily_max_fee:.2f}，按上限计费")
                            fee = daily_max_fee
                    else:
                        fee = 0.00
                        breakdown = ["无法计算：停车场无可用车位"]

        return JsonResponse({"success": True, "fee": f"{fee:.2f}", "breakdown": breakdown})

    except Exception as e:
        logger.exception("费率预览计算失败")
        return JsonResponse({"success": False, "message": f"计算失败: {str(e)}"}, status=500)


def _normalize_vehicle_type(vehicle_type: str) -> str:
    """将中英文车位类型转换为标准英文值"""
    vehicle_type = str(vehicle_type).strip().lower()
    mapping = {
        "全部": "all",
        "标准": "standard",
        "标准车位": "standard",
        "残疾人": "disabled",
        "残疾人车位": "disabled",
        "vip": "vip",
        "vip车位": "vip",
        "大型": "large",
        "大型车位": "large",
    }
    return mapping.get(vehicle_type, vehicle_type if vehicle_type in ["all", "standard", "disabled", "vip", "large"] else "all")


def _normalize_rate_type(rate_type: str) -> str:
    """将中英文包月/包年类型转换为标准英文值"""
    rate_type = str(rate_type).strip().lower()
    mapping = {
        "月卡": "month",
        "季卡": "quarter",
        "年卡": "year",
    }
    return mapping.get(rate_type, rate_type if rate_type in ["month", "quarter", "year"] else rate_type)


@staff_member_required
@require_http_methods(["GET"])
def pricing_template_download(request):
    """下载费率模板（压缩包，包含PDF说明和Excel模板）"""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "费率模板导入"

    # 定义样式
    # 标题样式
    title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")

    # 表头样式 - 不同类型使用不同颜色
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    basic_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 蓝色
    rule_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # 绿色
    month_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # 橙色
    overtime_header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")  # 深橙色
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 数据行样式
    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # 边框样式
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # 说明区域样式
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    note_font = Font(name="微软雅黑", size=9, italic=True, color="000000")

    # 冻结第一行
    ws.freeze_panes = "A2"

    # 第一行：标题
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "停车场费率模板导入表"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 30

    # 第二行：说明
    ws.merge_cells("A2:E2")
    note_cell = ws["A2"]
    note_cell.value = "📋 填写说明：请按照下方示例填写，每个模板之间用空行分隔"
    note_cell.font = note_font
    note_cell.fill = note_fill
    note_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 25

    # 第三行：模板基本信息表头
    row = 3
    basic_headers = [
        "模板名称 *",
        "模板描述",
        "免费时长(分钟) *",
        "每日收费上限(元)",
        "是否启用(是/否)",
    ]
    ws.append(basic_headers)
    for col_idx, cell in enumerate(ws[row], start=1):
        cell.fill = basic_header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[row].height = 35

    # 第四行：示例数据 - 模板基本信息
    row = 4
    example_basic = ["标准阶梯收费", "标准停车场费率模板，适用于普通停车场", "15", "100.00", "是"]
    ws.append(example_basic)
    for col_idx, cell in enumerate(ws[row], start=1):
        cell.font = data_font
        cell.border = thin_border
        if col_idx in [3, 4]:  # 数字列右对齐
            cell.alignment = number_alignment
        else:
            cell.alignment = data_alignment
    ws.row_dimensions[row].height = 25

    # 第五行：费率规则表头
    row = 5
    rule_headers = [
        "规则起始分钟 *",
        "规则结束分钟(空=无上限)",
        "每小时费率(元) *",
        "车位类型 *",
        "排序(数字越小越先) *",
    ]
    ws.append(rule_headers)
    for col_idx, cell in enumerate(ws[row], start=1):
        cell.fill = rule_header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[row].height = 35

    # 第六-八行：示例数据 - 费率规则（支持中英文）
    example_rules = [
        ["0", "60", "5.00", "全部", "0"],  # 使用中文示例，系统会自动转换
        ["60", "120", "8.00", "标准", "1"],
        ["120", "", "10.00", "VIP", "2"],
    ]
    for rule_data in example_rules:
        row += 1
        ws.append(rule_data)
        for col_idx, cell in enumerate(ws[row], start=1):
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [1, 2, 3, 5]:  # 数字列右对齐
                cell.alignment = number_alignment
            elif col_idx == 4:  # 车位类型居中
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
        ws.row_dimensions[row].height = 25

    # 第九行：包月/包年费率表头
    row = 9
    month_year_headers = [
        "包月/包年类型 *",
        "价格(元) *",
        "车位类型 *",
        "说明",
        "是否启用(是/否)",
    ]
    ws.append(month_year_headers)
    for col_idx, cell in enumerate(ws[row], start=1):
        cell.fill = month_header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[row].height = 35

    # 第十行：示例数据 - 包月/包年费率（支持中英文）
    row = 10
    example_month = ["月卡", "300.00", "全部", "包月优惠套餐", "是"]  # 使用中文示例，系统会自动转换
    ws.append(example_month)
    for col_idx, cell in enumerate(ws[row], start=1):
        cell.font = data_font
        cell.border = thin_border
        if col_idx == 2:  # 价格右对齐
            cell.alignment = number_alignment
        elif col_idx in [1, 3, 5]:  # 类型和启用状态居中
            cell.alignment = center_alignment
        else:
            cell.alignment = data_alignment
    ws.row_dimensions[row].height = 25

    # 第十一行：超时收费表头
    row = 11
    overtime_headers = [
        "超时费用(元/小时) *",
        "超时起始小时(默认24) *",
        "车位类型 *",
        "说明",
        "是否启用(是/否)",
    ]
    ws.append(overtime_headers)
    for col_idx, cell in enumerate(ws[row], start=1):
        cell.fill = overtime_header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[row].height = 35

    # 第十二行：示例数据 - 超时收费
    row = 12
    example_overtime = ["15.00", "24", "all", "超过24小时后的超时费用", "是"]
    ws.append(example_overtime)
    for col_idx, cell in enumerate(ws[row], start=1):
        cell.font = data_font
        cell.border = thin_border
        if col_idx in [1, 2]:  # 数字列右对齐
            cell.alignment = number_alignment
        elif col_idx in [3, 5]:  # 类型和启用状态居中
            cell.alignment = center_alignment
        else:
            cell.alignment = data_alignment
    ws.row_dimensions[row].height = 25

    # 添加详细说明区域
    row = 13
    ws.append([])  # 空行
    row = 14

    # 说明标题
    ws.merge_cells(f"A{row}:E{row}")
    note_title = ws[f"A{row}"]
    note_title.value = "📝 填写说明"
    note_title.font = Font(name="微软雅黑", size=12, bold=True, color="000000")
    note_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    note_title.alignment = Alignment(horizontal="left", vertical="center")
    note_title.border = thin_border
    ws.row_dimensions[row].height = 30

    # 详细说明
    notes = [
        "1. 模板基本信息（必填）：",
        "   • 模板名称：必填，唯一标识，如'标准阶梯收费'、'优惠收费'等",
        "   • 免费时长：必填，停车多长时间内免费（单位：分钟）",
        "   • 每日收费上限：选填，留空表示不设上限",
        "",
        "2. 费率规则（至少一条）：",
        "   • 起始分钟：必填，规则开始生效的分钟数（包含）",
        "   • 结束分钟：选填，留空表示无上限",
        "   • 车位类型：必填，可选值：",
        "     - 英文：all(全部)、standard(标准)、disabled(残疾人)、vip(VIP)、large(大型)",
        "     - 中文：全部、标准、残疾人、VIP、大型（系统会自动转换为英文）",
        "   • 排序：必填，数字越小越先执行，建议从0开始",
        "",
        "3. 包月/包年费率（可选）：",
        "   • 类型：必填，可选值：",
        "     - 英文：month(月卡)、quarter(季卡)、year(年卡)",
        "     - 中文：月卡、季卡、年卡（系统会自动转换为英文）",
        "   • 价格：必填，包月/包年费用（单位：元）",
        "",
        "4. 超时收费（可选）：",
        "   • 超时费用：必填，超过每日上限后的收费标准（单位：元/小时）",
        "   • 超时起始小时：必填，超过多少小时后开始收取超时费用（默认24小时）",
        "",
        "5. 注意事项：",
        "   • 每个模板之间用空行分隔",
        "   • 带 * 的字段为必填项",
        "   • 费率规则的时间段不能重叠",
        "   • 模板名称不能重复",
    ]

    for note in notes:
        row += 1
        ws.merge_cells(f"A{row}:E{row}")
        note_cell = ws[f"A{row}"]
        note_cell.value = note
        note_cell.font = note_font
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if note.startswith(("1.", "2.", "3.", "4.", "5.")):
            note_cell.font = Font(name="微软雅黑", size=9, bold=True, color="000000")
        ws.row_dimensions[row].height = 20 if note else 10

    # 自动调整列宽（自适应内容宽度）
    for col in range(1, 6):  # A到E列
        col_letter = get_column_letter(col)
        max_length = 0
        # 遍历该列的所有单元格，找到最大宽度
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    # 计算单元格内容长度（中文字符按2个字符宽度计算）
                    cell_value = str(cell.value)
                    length = sum(2 if ord(char) > 127 else 1 for char in cell_value)
                    max_length = max(max_length, length)
        
        # 设置列宽，最小10，最大50，加上一些边距
        width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = width

    # 设置打印区域
    ws.print_area = f"A1:E{row}"

    # 保存Excel到内存
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # 生成PDF使用说明
    pdf_content = _generate_pdf_manual()
    
    # 创建压缩包
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # 添加Excel模板
        zip_file.writestr("费率模板导入模板.xlsx", excel_buffer.getvalue())
        # 添加PDF使用说明
        zip_file.writestr("使用说明.pdf", pdf_content)
    
    zip_buffer.seek(0)
    
    # 生成响应
    response = HttpResponse(
        zip_buffer.getvalue(),
        content_type="application/zip"
    )
    response["Content-Disposition"] = 'attachment; filename="费率模板导入包.zip"'
    return response


def _generate_pdf_manual() -> bytes:
    """生成PDF使用说明文档"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib import colors
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 标题样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e293b'),
            spaceAfter=30,
            alignment=1,  # 居中
        )
        
        # 小标题样式
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#334155'),
            spaceAfter=12,
            spaceBefore=12,
        )
        
        # 正文样式
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#475569'),
            leading=16,
        )
        
        # 标题
        story.append(Paragraph("停车场费率模板导入使用说明", title_style))
        story.append(Spacer(1, 0.5*cm))
        
        # 概述
        story.append(Paragraph("一、概述", heading_style))
        story.append(Paragraph(
            "本系统支持通过Excel文件批量导入费率模板。每个Excel文件应包含一个费率模板的完整信息，"
            "包括模板基本信息、费率规则、包月/包年费率和超时收费等。",
            normal_style
        ))
        story.append(Spacer(1, 0.3*cm))
        
        # 模板基本信息
        story.append(Paragraph("二、模板基本信息（必填）", heading_style))
        story.append(Paragraph("• <b>模板名称</b>：必填，唯一标识，如'标准阶梯收费'、'优惠收费'等", normal_style))
        story.append(Paragraph("• <b>模板描述</b>：选填，模板的详细说明", normal_style))
        story.append(Paragraph("• <b>免费时长</b>：必填，停车多长时间内免费（单位：分钟）", normal_style))
        story.append(Paragraph("• <b>每日收费上限</b>：选填，留空表示不设上限（单位：元）", normal_style))
        story.append(Paragraph("• <b>是否启用</b>：选填，填写'是'或'否'，默认为'是'", normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # 费率规则
        story.append(Paragraph("三、费率规则（至少一条，必填）", heading_style))
        story.append(Paragraph("• <b>规则起始分钟</b>：必填，规则开始生效的分钟数（包含）", normal_style))
        story.append(Paragraph("• <b>规则结束分钟</b>：选填，留空表示无上限", normal_style))
        story.append(Paragraph("• <b>每小时费率</b>：必填，该时间段的收费标准（单位：元/小时）", normal_style))
        story.append(Paragraph("• <b>车位类型</b>：必填，可选值：", normal_style))
        story.append(Paragraph("  - 英文：all(全部)、standard(标准)、disabled(残疾人)、vip(VIP)、large(大型)", normal_style))
        story.append(Paragraph("  - 中文：全部、标准、残疾人、VIP、大型（系统会自动转换为英文）", normal_style))
        story.append(Paragraph("• <b>排序</b>：必填，数字越小越先执行，建议从0开始", normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # 包月/包年费率
        story.append(Paragraph("四、包月/包年费率（可选）", heading_style))
        story.append(Paragraph("• <b>类型</b>：必填，可选值：", normal_style))
        story.append(Paragraph("  - 英文：month(月卡)、quarter(季卡)、year(年卡)", normal_style))
        story.append(Paragraph("  - 中文：月卡、季卡、年卡（系统会自动转换为英文）", normal_style))
        story.append(Paragraph("• <b>价格</b>：必填，包月/包年费用（单位：元）", normal_style))
        story.append(Paragraph("• <b>车位类型</b>：必填，同费率规则中的车位类型", normal_style))
        story.append(Paragraph("• <b>说明</b>：选填，包月/包年套餐的详细说明", normal_style))
        story.append(Paragraph("• <b>是否启用</b>：选填，填写'是'或'否'，默认为'是'", normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # 超时收费
        story.append(Paragraph("五、超时收费（可选）", heading_style))
        story.append(Paragraph("• <b>超时费用</b>：必填，超过每日上限后的收费标准（单位：元/小时）", normal_style))
        story.append(Paragraph("• <b>超时起始小时</b>：必填，超过多少小时后开始收取超时费用（默认24小时）", normal_style))
        story.append(Paragraph("• <b>车位类型</b>：必填，同费率规则中的车位类型", normal_style))
        story.append(Paragraph("• <b>说明</b>：选填，超时收费的详细说明", normal_style))
        story.append(Paragraph("• <b>是否启用</b>：选填，填写'是'或'否'，默认为'是'", normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # 注意事项
        story.append(Paragraph("六、注意事项", heading_style))
        story.append(Paragraph("• 每个Excel文件只包含一个费率模板", normal_style))
        story.append(Paragraph("• 带 * 的字段为必填项", normal_style))
        story.append(Paragraph("• 费率规则的时间段不能重叠", normal_style))
        story.append(Paragraph("• 如果模板名称已存在，系统会自动重命名（添加序号）", normal_style))
        story.append(Paragraph("• 支持中英文混合输入，系统会自动转换", normal_style))
        story.append(Paragraph("• 建议先下载模板文件，参考示例填写", normal_style))
        story.append(Spacer(1, 0.3*cm))
        
        # 构建PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    except ImportError:
        # 如果没有reportlab，生成简单的文本说明
        text_content = """
停车场费率模板导入使用说明

一、概述
本系统支持通过Excel文件批量导入费率模板。每个Excel文件应包含一个费率模板的完整信息。

二、模板基本信息（必填）
• 模板名称：必填，唯一标识
• 免费时长：必填，单位：分钟
• 每日收费上限：选填，单位：元

三、费率规则（至少一条）
• 车位类型支持中英文：
  - 英文：all, standard, disabled, vip, large
  - 中文：全部、标准、残疾人、VIP、大型

四、包月/包年费率（可选）
• 类型支持中英文：
  - 英文：month, quarter, year
  - 中文：月卡、季卡、年卡

五、注意事项
• 每个Excel文件只包含一个费率模板
• 如果模板名称已存在，系统会自动重命名
• 支持中英文混合输入，系统会自动转换
"""
        return text_content.encode('utf-8')


@staff_member_required
@require_http_methods(["POST"])
def pricing_template_import(request):
    """
    批量导入费率模板（Excel）
    
    支持一次上传多个Excel文件，每个文件包含一个费率模板。
    """
    # 获取所有上传的文件
    files = request.FILES.getlist("files")
    
    if not files:
        return JsonResponse({"success": False, "message": "请选择要上传的文件"}, status=400)

    success_count = 0
    error_count = 0
    errors = []

    # 处理每个文件
    for file_idx, file in enumerate(files, start=1):
        file_name = file.name
        
        # 验证文件类型
        if not file.name.endswith((".xlsx", ".xls")):
            error_count += 1
            errors.append(f"文件 '{file_name}': 只支持Excel文件（.xlsx, .xls）")
            continue

        try:
            # 读取Excel文件
            wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
            ws = wb.active

            # 识别数据开始行（跳过标题、说明和表头）
            header_keywords = ["模板名称", "规则起始分钟", "包月/包年类型", "超时费用", "填写说明", "费率模板导入表"]
            skip_keywords = ["📋", "📝", "说明：", "注意事项：", "1.", "2.", "3.", "4.", "5."]
            
            data_start_row = 1
            for row_idx in range(1, min(20, ws.max_row + 1)):  # 检查前20行
                row_values = [cell.value for cell in ws[row_idx]]
                if not any(row_values):
                    continue
                    
                first_col = str(row_values[0]).strip() if row_values[0] else ""
                
                # 如果第一列包含表头关键词，跳过
                if any(keyword in first_col for keyword in header_keywords):
                    data_start_row = row_idx + 1
                    continue
                
                # 如果第一列是说明关键词，跳过
                if any(first_col.startswith(keyword) for keyword in skip_keywords):
                    data_start_row = row_idx + 1
                    continue
                
                # 如果第一列是实际数据（不是表头），从这里开始
                if first_col and first_col not in ["month", "quarter", "year"]:
                    # 检查是否是数字或模板名称格式（不是表头）
                    is_number = first_col.replace(".", "").isdigit()
                    is_template_name = (
                        len(first_col) > 0 
                        and not any(keyword in first_col for keyword in header_keywords)
                        and not any(first_col.startswith(kw) for kw in skip_keywords)
                    )
                    if is_number or is_template_name:
                        data_start_row = row_idx
                        break

            # 解析模板数据（每个文件只包含一个模板）
            template_data = None
            rules = []
            month_year_rates = []
            overtime_rates = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                if not any(row):
                    # 空行，跳过
                    continue

                # 跳过表头行和说明行
                first_col = str(row[0]).strip() if row[0] else ""
                
                # 跳过包含表头关键词的行
                if any(keyword in first_col for keyword in header_keywords):
                    continue
                
                # 跳过说明行
                if any(first_col.startswith(keyword) for keyword in skip_keywords):
                    continue

                # 判断行类型（根据第一列的值）
                if first_col and first_col not in ["month", "quarter", "year"] and not first_col.replace(".", "").isdigit() and first_col not in ["", "None"]:
                    # 模板基本信息行
                    template_name = str(row[0]).strip() if row[0] else ""
                    if not _validate_template_name(template_name):
                        # 跳过说明行和无效数据
                        continue
                    
                    try:
                        # 安全地解析free_minutes
                        free_minutes = 15
                        if len(row) > 2 and row[2]:
                            try:
                                free_minutes = int(row[2])
                                if free_minutes < 0 or free_minutes > 1440:
                                    errors.append(f"文件 '{file_name}' 第{row_idx}行：免费时长必须在0-1440分钟之间")
                                    continue
                            except (ValueError, TypeError):
                                free_minutes_str = str(row[2]).strip()
                                if free_minutes_str.isdigit():
                                    free_minutes = int(free_minutes_str)
                                    if free_minutes < 0 or free_minutes > 1440:
                                        errors.append(f"文件 '{file_name}' 第{row_idx}行：免费时长必须在0-1440分钟之间")
                                        continue
                                else:
                                    continue
                        
                        # 安全地解析daily_max_fee
                        daily_max_fee = None
                        if len(row) > 3 and row[3] and str(row[3]).strip() not in ["None", "", "每日收费上限(元)", "每日收费上限"]:
                            try:
                                daily_max_fee = float(row[3])
                                if daily_max_fee <= 0:
                                    daily_max_fee = None
                            except (ValueError, TypeError):
                                daily_max_fee = None
                        
                        template_data = {
                            "name": template_name,
                            "description": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                            "free_minutes": free_minutes,
                            "daily_max_fee": daily_max_fee,
                            "is_active": (
                                str(row[4]).strip().lower() in ["是", "yes", "true", "1"]
                                if len(row) > 4 and row[4]
                                else True
                            ),
                        }
                    except Exception as e:
                        errors.append(f"文件 '{file_name}' 第{row_idx}行模板基本信息解析错误: {str(e)}")
                        continue
                elif first_col.replace(".", "").isdigit() or first_col == "":
                    # 费率规则行（第一列是数字或空）
                    if len(row) >= 3 and row[0] is not None:
                        try:
                            start_minutes = int(row[0]) if row[0] else 0
                            if start_minutes < 0:
                                continue
                            
                            rate_per_hour = float(row[2]) if len(row) > 2 and row[2] else 0.0
                            if rate_per_hour <= 0:
                                continue
                            
                            end_minutes = None
                            if len(row) > 1 and row[1] and str(row[1]).strip():
                                try:
                                    end_minutes = int(row[1])
                                    if end_minutes <= start_minutes:
                                        continue
                                except (ValueError, TypeError):
                                    pass
                            
                            vehicle_type = str(row[3]).strip() if len(row) > 3 and row[3] else "all"
                            vehicle_type = _normalize_vehicle_type(vehicle_type)
                            
                            rule_data = {
                                "start_minutes": start_minutes,
                                "end_minutes": end_minutes,
                                "rate_per_hour": rate_per_hour,
                                "vehicle_type": vehicle_type,
                                "order": len(rules),
                            }
                            rules.append(rule_data)
                        except (ValueError, TypeError):
                            continue
                elif first_col in ["month", "quarter", "year"] or first_col in ["月卡", "季卡", "年卡"]:
                    # 包月/包年费率行
                    try:
                        rate_type_str = str(row[0]).strip() if row[0] else ""
                        rate_type = _normalize_rate_type(rate_type_str)
                        if rate_type not in ["month", "quarter", "year"]:
                            continue
                        
                        price = float(row[1]) if len(row) > 1 and row[1] else 0.0
                        if price <= 0:
                            continue
                        
                        vehicle_type = str(row[2]).strip() if len(row) > 2 and row[2] else "all"
                        vehicle_type = _normalize_vehicle_type(vehicle_type)
                        
                        month_year_data = {
                            "rate_type": rate_type,
                            "price": price,
                            "vehicle_type": vehicle_type,
                            "description": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                            "is_active": (
                                str(row[4]).strip().lower() in ["是", "yes", "true", "1"]
                                if len(row) > 4 and row[4]
                                else True
                            ),
                        }
                        month_year_rates.append(month_year_data)
                    except (ValueError, TypeError):
                        continue
                else:
                    # 可能是超时收费行（第一列是数字，表示超时费用）
                    try:
                        if not first_col.replace(".", "").isdigit():
                            continue
                        
                        overtime_fee = float(row[0]) if row[0] else 0.0
                        if overtime_fee <= 0:
                            continue
                        
                        overtime_start_hours = 24
                        if len(row) > 1 and row[1]:
                            try:
                                overtime_start_hours = int(row[1])
                                if overtime_start_hours < 1:
                                    continue
                            except (ValueError, TypeError):
                                pass
                        
                        vehicle_type = str(row[2]).strip() if len(row) > 2 and row[2] else "all"
                        valid_vehicle_types = ["all", "standard", "disabled", "vip", "large"]
                        if vehicle_type not in valid_vehicle_types:
                            vehicle_type = "all"
                        
                        overtime_data = {
                            "overtime_fee": overtime_fee,
                            "overtime_start_hours": overtime_start_hours,
                            "vehicle_type": vehicle_type,
                            "description": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                            "is_active": (
                                str(row[4]).strip().lower() in ["是", "yes", "true", "1"]
                                if len(row) > 4 and row[4]
                                else True
                            ),
                        }
                        overtime_rates.append(overtime_data)
                    except (ValueError, TypeError):
                        continue

            # 处理当前文件的模板
            if template_data:
                template_name = template_data.get("name", "").strip()
                if not _validate_template_name(template_name):
                    error_count += 1
                    errors.append(f"文件 '{file_name}': 模板名称无效或为说明行: '{template_name}'")
                elif not rules:
                    error_count += 1
                    errors.append(f"文件 '{file_name}': 模板 '{template_name}' 必须包含至少一条费率规则")
                else:
                    try:
                        template = _create_template_from_data(
                            template_data,
                            rules,
                            month_year_rates,
                            overtime_rates,
                        )
                        success_count += 1
                        logger.info(f"成功导入费率模板: {template.name} (文件: {file_name})")
                    except Exception as e:
                        error_count += 1
                        errors.append(f"文件 '{file_name}': 模板 '{template_name}': {str(e)}")
                        logger.exception(f"导入费率模板失败: {template_name} (文件: {file_name})")
            else:
                error_count += 1
                errors.append(f"文件 '{file_name}': 未找到有效的模板数据")

        except Exception as e:
            error_count += 1
            errors.append(f"文件 '{file_name}': 解析失败 - {str(e)}")
            logger.exception(f"Excel文件解析失败: {file_name}")

    return JsonResponse(
        {
            "success": True,
            "message": f"导入完成：成功 {success_count} 个，失败 {error_count} 个",
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors[:50],  # 最多返回50个错误
        }
    )


def _validate_template_name(name: str) -> bool:
    """
    验证模板名称是否有效
    
    过滤掉说明行和无效数据：
    - 不能以"·"开头（说明行标记）
    - 不能包含说明性关键词（必填、选填、单位、建议等）
    - 长度应该在合理范围内（3-100字符）
    - 不能是纯数字
    - 不能包含过多的标点符号
    
    Returns:
        bool: True表示有效，False表示无效
    """
    if not name or not isinstance(name, str):
        return False
    
    name = name.strip()
    
    # 长度检查
    if len(name) < 3 or len(name) > 100:
        return False
    
    # 不能以说明标记开头
    if name.startswith("·") or name.startswith("*") or name.startswith("-"):
        return False
    
    # 不能包含说明性关键词
    invalid_keywords = [
        "必填",
        "选填",
        "单位",
        "建议",
        "说明：",
        "注意事项：",
        "填写说明",
        "费率规则的时间段",
        "不能重叠",
        "数字越小",
        "留空表示",
        "不设上限",
        "包月/包年费用",
        "超时费用",
        "超时起始小时",
        "规则起始分钟",
        "规则结束分钟",
        "每小时费率",
        "模板名称",
        "模板描述",
        "免费时长",
        "每日收费上限",
        "是否启用",
    ]
    
    for keyword in invalid_keywords:
        if keyword in name:
            return False
    
    # 不能是纯数字
    if name.replace(".", "").isdigit():
        return False
    
    # 不能包含过多的标点符号（超过3个）
    punctuation_count = sum(1 for c in name if c in "，。、；：！？""''（）【】《》")
    if punctuation_count > 3:
        return False
    
    # 不能是表头格式（包含括号和星号）
    if "(" in name and ")" in name and "*" in name:
        return False
    
    return True


def _create_template_from_data(
    template_data: dict,
    rules: list[dict],
    month_year_rates: list[dict],
    overtime_rates: list[dict],
) -> PricingTemplate:
    """从数据字典创建费率模板"""
    # 验证模板名称
    template_name = template_data.get("name", "").strip()
    if not _validate_template_name(template_name):
        raise ValueError(f"模板名称无效或为说明行: '{template_name}'")
    
    # 检查模板名称是否已存在，如果存在则自动重命名
    original_name = template_name
    counter = 1
    while PricingTemplate.objects.filter(name=template_name).exists():
        template_name = f"{original_name} ({counter})"
        counter += 1
    
    # 如果名称被修改，更新template_data
    if template_name != original_name:
        template_data = template_data.copy()
        template_data["name"] = template_name
        logger.info(f"模板名称 '{original_name}' 已存在，自动重命名为 '{template_name}'")
    
    # 验证数据完整性
    # 必须有至少一条费率规则
    if not rules:
        raise ValueError(f"模板 '{template_name}' 必须包含至少一条费率规则")
    
    # 验证费率规则的有效性
    for rule in rules:
        if rule.get("rate_per_hour", 0) <= 0:
            raise ValueError(f"模板 '{template_name}' 的费率规则费率必须大于0")

    # 创建模板
    template = PricingTemplate.objects.create(
        name=template_name,
        description=template_data.get("description", ""),
        free_minutes=template_data.get("free_minutes", 15),
        daily_max_fee=template_data.get("daily_max_fee"),
        is_active=template_data.get("is_active", True),
    )

    # 创建费率规则
    for rule_data in rules:
        PricingRule.objects.create(
            template=template,
            start_minutes=rule_data["start_minutes"],
            end_minutes=rule_data.get("end_minutes"),
            rate_per_hour=rule_data["rate_per_hour"],
            vehicle_type=rule_data.get("vehicle_type", "all"),
            order=rule_data.get("order", 0),
        )

    # 创建包月/包年费率
    for rate_data in month_year_rates:
        MonthYearRate.objects.create(
            template=template,
            rate_type=rate_data["rate_type"],
            price=rate_data["price"],
            vehicle_type=rate_data.get("vehicle_type", "all"),
            description=rate_data.get("description", ""),
            is_active=rate_data.get("is_active", True),
        )

    # 创建超时收费
    for overtime_data in overtime_rates:
        OvertimeRate.objects.create(
            template=template,
            overtime_fee=overtime_data["overtime_fee"],
            overtime_start_hours=overtime_data.get("overtime_start_hours", 24),
            vehicle_type=overtime_data.get("vehicle_type", "all"),
            description=overtime_data.get("description", ""),
            is_active=overtime_data.get("is_active", True),
        )

    return template
