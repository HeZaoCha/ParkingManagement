"""
车位号批量创建服务

支持通过文档上传、范围设定等方式批量创建停车位。

Author: HeZaoCha
Created: 2025-12-11
Version: 1.1.0
"""

import re
from io import BytesIO

from django.core.exceptions import ValidationError
from loguru import logger
from openpyxl import Workbook, load_workbook

from .models import ParkingLot, ParkingSpace


class SpaceNumberParser:
    """车位号解析器"""

    @staticmethod
    def parse_range(start: str, end: str) -> tuple[list[str], bool]:
        """
        解析车位号范围

        Args:
            start: 起始车位号，如"A001"
            end: 结束车位号，如"A1000"

        Returns:
            tuple[list[str], bool]: (车位号列表, 是否成功)

        Examples:
            >>> parse_range("A001", "A1000")
            (["A001", "A002", ..., "A1000"], True)

            >>> parse_range("A001", "B001")
            ([], False)  # 无法确定范围
        """
        start = start.strip().upper()
        end = end.strip().upper()

        # 提取前缀和数字部分
        start_match = re.match(r"^([A-Z]+)(\d+)$", start)
        end_match = re.match(r"^([A-Z]+)(\d+)$", end)

        if not start_match or not end_match:
            return [], False

        start_prefix = start_match.group(1)
        start_num = int(start_match.group(2))
        end_prefix = end_match.group(1)
        end_num = int(end_match.group(2))

        # 前缀必须相同
        if start_prefix != end_prefix:
            return [], False

        # 结束数字必须大于起始数字
        if end_num < start_num:
            return [], False

        # 生成车位号列表（使用生成器表达式优化内存）
        num_width = len(start_match.group(2))  # 保持数字位数

        # Python 3.13优化：使用列表推导式，更高效
        space_numbers = [
            f"{start_prefix}{str(num).zfill(num_width)}" for num in range(start_num, end_num + 1)
        ]

        return space_numbers, True

    @staticmethod
    def parse_from_text(text: str) -> list[str]:
        """
        从文本中解析车位号

        支持多种格式：
        - 每行一个：A001\nA002\nA003
        - 逗号分隔：A001,A002,A003
        - 范围：A001-A100

        Args:
            text: 文本内容

        Returns:
            list[str]: 车位号列表
        """
        space_numbers = []
        lines = text.strip().split("\n")

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # 处理逗号分隔
            if "," in line:
                parts = [p.strip() for p in line.split(",")]
                space_numbers.extend(parts)
                continue

            # 处理范围（A001-A100）
            if "-" in line:
                parts = line.split("-", 1)
                if len(parts) == 2:
                    start, end = parts[0].strip(), parts[1].strip()
                    numbers, success = SpaceNumberParser.parse_range(start, end)
                    if success:
                        space_numbers.extend(numbers)
                    else:
                        logger.warning(f"无法解析范围: {line}")
                continue

            # 单个车位号
            space_numbers.append(line)

        return space_numbers

    @staticmethod
    def parse_from_excel(
        file_content: bytes, sheet_name: str | None = None
    ) -> list[dict[str, str]]:
        """
        从Excel文件中解析车位号

        Excel格式：
        - 第一列：车位号
        - 第二列：楼层（可选）
        - 第三列：区域（可选）
        - 第四列：车位类型（可选）

        Args:
            file_content: Excel文件内容（字节）
            sheet_name: 工作表名称，None表示使用第一个工作表

        Returns:
            list[dict[str, str]]: 车位信息列表，格式：
                [{"space_number": "A001", "floor": "B2", "area": "A区", "space_type": "standard"}, ...]
        """
        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            spaces = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                space_number = str(row[0]).strip() if row[0] else None
                if not space_number:
                    continue

                space_info = {
                    "space_number": space_number,
                    "floor": str(row[1]).strip() if len(row) > 1 and row[1] else None,
                    "area": str(row[2]).strip() if len(row) > 2 and row[2] else None,
                    "space_type": str(row[3]).strip() if len(row) > 3 and row[3] else "standard",
                }

                spaces.append(space_info)

            return spaces

        except Exception as e:
            logger.exception("解析Excel文件失败")
            raise ValidationError(f"Excel文件解析失败: {str(e)}")

    @staticmethod
    def generate_excel_template() -> bytes:
        """
        生成车位号Excel模板

        Returns:
            bytes: Excel文件内容
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "车位号模板"

        # 表头
        headers = ["车位号", "楼层", "区域", "车位类型"]
        ws.append(headers)

        # 示例数据
        examples = [
            ["A001", "B2", "A区", "standard"],
            ["A002", "B2", "A区", "standard"],
            ["B001", "B2", "B区", "vip"],
        ]
        for row in examples:
            ws.append(row)

        # 添加说明
        ws.append([])
        ws.append(["说明："])
        ws.append(["1. 车位号：必填，如A001、B2-A001等"])
        ws.append(["2. 楼层：选填，如B2、1F等"])
        ws.append(["3. 区域：选填，如A区、B区等"])
        ws.append(["4. 车位类型：选填，standard/vip/large/disabled"])

        # 调整列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15

        # 生成文件
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class SpaceCreationService:
    """车位创建服务"""

    @staticmethod
    def create_spaces_from_range(
        parking_lot: ParkingLot,
        start: str,
        end: str,
        space_type: str = "standard",
        floor: str | None = None,
        area: str | None = None,
    ) -> tuple[int, int, list[str], list[str], bool, str]:
        """
        从范围创建车位（增强版本，返回详细结果）

        Args:
            parking_lot: 停车场对象
            start: 起始车位号
            end: 结束车位号
            space_type: 车位类型
            floor: 楼层
            area: 区域

        Returns:
            Tuple[int, int, List[str], List[str], bool, str]:
            (创建数量, 跳过数量, 创建成功的车位号列表, 跳过的车位号列表, 是否成功, 错误消息)
        """
        space_numbers, success = SpaceNumberParser.parse_range(start, end)

        if not success:
            return 0, 0, [], [], False, "无法解析车位号范围，请使用模板上传方式"

        if not space_numbers:
            return 0, 0, [], [], False, "未生成任何车位号"

        # 检查是否已存在
        existing = set(
            ParkingSpace.objects.filter(
                parking_lot=parking_lot, space_number__in=space_numbers
            ).values_list("space_number", flat=True)
        )

        # 分离新创建和已存在的
        spaces_to_create = []
        skipped_numbers = []

        for space_number in space_numbers:
            if space_number in existing:
                skipped_numbers.append(space_number)
            else:
                spaces_to_create.append(
                    ParkingSpace(
                        parking_lot=parking_lot,
                        space_number=space_number,
                        space_type=space_type,
                        floor=floor,
                        area=area,
                        is_occupied=False,
                        is_reserved=False,
                    )
                )

        # 批量创建
        created_numbers = []
        if spaces_to_create:
            ParkingSpace.objects.bulk_create(spaces_to_create)
            created_numbers = [s.space_number for s in spaces_to_create]

        # 更新停车场总车位数
        parking_lot.total_spaces = ParkingSpace.objects.filter(parking_lot=parking_lot).count()
        parking_lot.save(update_fields=["total_spaces"])

        return (
            len(created_numbers),
            len(skipped_numbers),
            created_numbers,
            skipped_numbers,
            True,
            "创建完成",
        )

    @staticmethod
    def create_spaces_from_file(
        parking_lot: ParkingLot, file_content: bytes, file_type: str
    ) -> tuple[int, int, list[str], list[str], list[dict[str, str]], bool, str]:
        """
        从文件创建车位（增强版本，返回详细结果）

        Args:
            parking_lot: 停车场对象
            file_content: 文件内容（字节）
            file_type: 文件类型（txt/md/xlsx）

        Returns:
            Tuple[int, int, List[str], List[str], List[dict], bool, str]:
            (创建数量, 跳过数量, 创建成功的车位号列表, 跳过的车位号列表, 解析失败的行信息, 是否成功, 错误消息)
        """
        created_numbers = []
        skipped_numbers = []
        failed_lines = []

        try:
            # 使用match/case优化文件类型处理（Python 3.10+特性）
            match file_type:
                case "txt" | "md":
                    # 文本文件
                    text = file_content.decode("utf-8")
                    lines = text.strip().split("\n")
                    space_numbers = []

                    for line_num, line in enumerate(lines, 1):
                        line = line.strip()
                        if not line or line.startswith("#"):
                            continue
                        parsed = SpaceNumberParser.parse_from_text(line)
                        if parsed:
                            space_numbers.extend(parsed)
                        else:
                            failed_lines.append(
                                {"line": line_num, "content": line, "reason": "无法解析车位号格式"}
                            )

                    # 检查已存在的车位
                    existing = set(
                        ParkingSpace.objects.filter(
                            parking_lot=parking_lot, space_number__in=space_numbers
                        ).values_list("space_number", flat=True)
                    )

                    # 使用生成器表达式优化内存使用（Python 3.13推荐）
                    spaces_to_create = [
                        ParkingSpace(
                            parking_lot=parking_lot,
                            space_number=space_number,
                            is_occupied=False,
                            is_reserved=False,
                        )
                        for space_number in space_numbers
                        if space_number not in existing
                    ]
                    # 收集跳过的车位号
                    skipped_numbers.extend(
                        space_number for space_number in space_numbers if space_number in existing
                    )

                    if spaces_to_create:
                        ParkingSpace.objects.bulk_create(spaces_to_create)
                        created_numbers = [s.space_number for s in spaces_to_create]

                    # 更新总车位数
                    parking_lot.total_spaces = ParkingSpace.objects.filter(
                        parking_lot=parking_lot
                    ).count()
                    parking_lot.save(update_fields=["total_spaces"])

                    return (
                        len(created_numbers),
                        len(skipped_numbers),
                        created_numbers,
                        skipped_numbers,
                        failed_lines,
                        True,
                        "创建完成",
                    )

                case "xlsx":
                    # Excel文件
                    spaces_info = SpaceNumberParser.parse_from_excel(file_content)

                    # 检查已存在的车位
                    existing = set(
                        ParkingSpace.objects.filter(
                            parking_lot=parking_lot,
                            space_number__in=[info["space_number"] for info in spaces_info],
                        ).values_list("space_number", flat=True)
                    )

                    # 使用生成器表达式优化内存使用（Python 3.13推荐）
                    spaces_to_create = [
                        ParkingSpace(
                            parking_lot=parking_lot,
                            space_number=info["space_number"],
                            floor=info.get("floor"),
                            area=info.get("area"),
                            space_type=info.get("space_type", "standard"),
                            is_occupied=False,
                            is_reserved=False,
                        )
                        for info in spaces_info
                        if info["space_number"] not in existing
                    ]
                    # 收集跳过的车位号
                    skipped_numbers.extend(
                        info["space_number"]
                        for info in spaces_info
                        if info["space_number"] in existing
                    )

                    if spaces_to_create:
                        ParkingSpace.objects.bulk_create(spaces_to_create)
                        created_numbers = [s.space_number for s in spaces_to_create]

                    # 更新总车位数
                    parking_lot.total_spaces = ParkingSpace.objects.filter(
                        parking_lot=parking_lot
                    ).count()
                    parking_lot.save(update_fields=["total_spaces"])

                    return (
                        len(created_numbers),
                        len(skipped_numbers),
                        created_numbers,
                        skipped_numbers,
                        failed_lines,
                        True,
                        "创建完成",
                    )

                case _:
                    return 0, 0, [], [], [], False, f"不支持的文件类型: {file_type}"

        except Exception as e:
            logger.exception("从文件创建车位失败")
            return 0, 0, [], [], [], False, f"创建失败: {str(e)}"

    @staticmethod
    def create_spaces_from_file_simple(
        parking_lot: ParkingLot, file_content: bytes, file_type: str
    ) -> tuple[int, bool, str]:
        """
        从文件创建车位（简化版本，保持向后兼容）
        """
        created, skipped, created_list, skipped_list, failed_lines, success, message = (
            SpaceCreationService.create_spaces_from_file(parking_lot, file_content, file_type)
        )
        if not success:
            return 0, False, message
        msg = f"成功创建 {created} 个车位"
        if skipped > 0:
            msg += f"，跳过 {skipped} 个已存在的车位"
        if failed_lines:
            msg += f"，{len(failed_lines)} 行解析失败"
        return created, True, msg
